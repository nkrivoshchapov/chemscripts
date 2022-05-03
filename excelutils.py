from openpyxl import load_workbook, Workbook
import openpyxl.utils as opxutils
import xlcalculator as xlc


class ExcelSheet:
    def __init__(self):
        self.datablocks = []
        self.blocknames = []

    def block(self, name):
        return self.datablocks[self.blocknames.index(name)]

    def add_key(self, blockname, key, position=0):
        self.block(blockname)['keys'].insert(position, key)

    def remove_key(self, blockname, key):
        block = self.block(blockname)
        for item in block['data']:
            del item[key]
        block['keys'].remove(key)

    def add_block(self, blockname, cols=None):
        assert blockname not in self.blocknames
        newblock = {'name': blockname, 'data': []}
        if cols is None:
            newblock['keys'] = []
        else:
            newblock['keys'] = cols
        self.datablocks.append(newblock)
        self.blocknames.append(blockname)

    def add_row(self, blockname, data):
        current_block = self.block(blockname)
        for key in data.keys():
            assert key in current_block['keys']

        newitem = {}
        for key in current_block['keys']:
            if key in data:
                newitem[key] = data[key]
            else:
                newitem[key] = None
        current_block['data'].append(newitem)

    def save_xlsx(self, filename, oldfile=None):
        if oldfile is None:
            wb = Workbook()
            ws = wb.active
            sheetname = "Main"
            ws.title = sheetname
        else:
            wb = load_workbook(filename=oldfile)
            ws = wb.active

        i = 1
        for datablock in self.datablocks:
            if len(datablock["data"]) != 0:
                ws.cell(i, 1).value = datablock['name']
                i += 1

                for j, key in enumerate(datablock['keys']):
                    ws.cell(i, j + 1).value = key
                i += 1

                for dataitem in datablock['data']:
                    for j, key in enumerate(datablock['keys']):
                        if dataitem[key] is not None:
                            ws.cell(i, j + 1).value = dataitem[key]
                    i += 1
                i += 1  # One empty row between sections
        wb.save(filename)

    def read_xlsx(self, filename, get_values=False):
        self.datablocks = []

        if get_values:
            compiler = xlc.ModelCompiler()
            new_model = compiler.read_and_parse_archive(filename)
            evaluator = xlc.Evaluator(new_model)
        workbook = load_workbook(filename=filename)
        ws = workbook.active
        sheetname = workbook.active.title
        vsize = ws.max_row
        hsize = ws.max_column

        section_separators = []
        section_names = []
        for i in range(1, vsize+2):
            if i - 1 in section_separators and ws.cell(i, 1).value is None:
                break
            elif ws.cell(i, 1).value is None:
                section_separators.append(i)
            elif i == 1 or i - 1 in section_separators:
                section_names.append(ws.cell(i, 1).value)

        assert len(section_separators) == len(section_names)
        for i in range(len(section_separators)):
            section_name = section_names[i]
            if i == 0:
                section_start = 2
            else:
                section_start = section_separators[i - 1] + 2
            section_end = section_separators[i]

            keys = []
            for j in range(hsize):
                value = ws.cell(section_start, j + 1).value
                if value is not None:
                    keys.append(value)
                else:
                    break

            data = []
            cells = []
            for i in range(section_start + 1, section_end):
                newitem = {}
                newcellrow = {}
                for key_idx, key in enumerate(keys):
                    cellname = ExcelSheet.get_cell_name(i, key_idx + 1, sheetname)
                    newcellrow[key] = cellname
                    if ws.cell(i, key_idx + 1).value is not None:
                        if get_values:
                            newitem[key] = evaluator.evaluate(cellname).value
                        else:
                            newitem[key] = ws.cell(i, key_idx + 1).value
                    else:
                        newitem[key] = None
                data.append(newitem)
                cells.append(newcellrow)

            newsection = {'name': section_name,
                          'keys': keys,
                          'data': data,
                          'cells': cells}
            self.datablocks.append(newsection)
            self.blocknames.append(section_name)

    @staticmethod
    def get_cell_name(row, col, sheetname):
        return "%s!%s%d" % (sheetname, opxutils.get_column_letter(col), row)
